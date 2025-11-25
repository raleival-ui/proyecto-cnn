import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from app.config import Config
from app.utils.metrics_calculator import MetricsCalculator

class ReportGenerator:
    def __init__(self):
        self.config = Config()
        self.metrics_calc = MetricsCalculator()
    
    def create_enhanced_results(self, results):
        """Crear resultados mejorados con diagnóstico y clasificación"""
        enhanced_results = []
        
        for result in results:
            if result['Prediccion'] != 'ERROR':
                diagnosis = self.metrics_calc.extract_diagnosis_from_filename(result['Nombre_Archivo'])
                
                classification_result = self.metrics_calc.calculate_classification_result(
                    result['Prediccion'], diagnosis
                )
                
                enhanced_result = result.copy()
                enhanced_result['Diagnostico'] = diagnosis
                enhanced_result['Resultado'] = classification_result
                
                enhanced_results.append(enhanced_result)
            else:
                enhanced_result = result.copy()
                enhanced_result['Diagnostico'] = 'ERROR'
                enhanced_result['Resultado'] = 'ERROR'
                enhanced_results.append(enhanced_result)
        
        return enhanced_results
    
    def create_dataframe(self, results):
        """Crear DataFrame con los resultados"""
        df = pd.DataFrame(results)
        df['Fecha_Procesamiento'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        columns_order = ['Nombre_Archivo', 'Prediccion', 'Diagnostico', 'Resultado', 'Confianza', 
                        'Prob_Benign', 'Prob_Malignant', 'Prob_Normal', 'Fecha_Procesamiento']
        available_columns = [col for col in columns_order if col in df.columns]
        df = df[available_columns]
        
        return df
    
    def create_excel_report(self, results):
        """Crear reporte Excel con formato profesional"""
        df = self.create_dataframe(results)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados Análisis"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        colors = {
            'VP': PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
            'VN': PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
            'FP': PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            'FN': PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
            'ERROR': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        }
        
        result_fonts = {
            'VP': Font(color="155724", bold=True),
            'VN': Font(color="0C5460", bold=True),
            'FP': Font(color="856404", bold=True),
            'FN': Font(color="721C24", bold=True),
            'ERROR': Font(color="6C757D", bold=True)
        }
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = df.columns.tolist()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_num, row_data in enumerate(df.iterrows(), 2):
            _, data = row_data
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if col_num == 4 and 'Resultado' in df.columns:
                    result_value = str(value)
                    if result_value in colors:
                        cell.fill = colors[result_value]
                        cell.font = result_fonts[result_value]
                
                if col_num in [6, 7, 8]:
                    try:
                        prob_value = float(value)
                        cell.value = f"{prob_value:.4f}"
                        if prob_value > 0.7:
                            cell.font = Font(color="155724", bold=True)
                        elif prob_value > 0.5:
                            cell.font = Font(color="856404", bold=True)
                    except:
                        pass
                
                if col_num == 5:
                    try:
                        conf_value = float(value)
                        cell.value = f"{conf_value:.4f}"
                        if conf_value > 0.8:
                            cell.font = Font(color="155724", bold=True)
                        elif conf_value > 0.6:
                            cell.font = Font(color="856404", bold=True)
                        else:
                            cell.font = Font(color="721C24", bold=True)
                    except:
                        pass
        
        column_widths = {
            'A': 25, 'B': 15, 'C': 15, 'D': 12, 'E': 12,
            'F': 15, 'G': 15, 'H': 15, 'I': 20
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        self._create_summary_sheet(wb, df)
        
        return self._excel_to_bytes(wb)
    
    def _create_summary_sheet(self, wb, df):
        """Crear hoja de resumen con métricas"""
        summary_ws = wb.create_sheet("Resumen Métricas")
        
        if 'Resultado' in df.columns:
            counts = df['Resultado'].value_counts()
            vp = counts.get('VP', 0)
            vn = counts.get('VN', 0)
            fp = counts.get('FP', 0)
            fn = counts.get('FN', 0)
            total = vp + vn + fp + fn
            
            if total > 0:
                precision = vp / (vp + fp) if (vp + fp) > 0 else 0
                sensitivity = vp / (vp + fn) if (vp + fn) > 0 else 0
                specificity = vn / (vn + fp) if (vn + fp) > 0 else 0
                f1_score = 2 * (precision * sensitivity) / (precision + sensitivity) if (precision + sensitivity) > 0 else 0
                accuracy = (vp + vn) / total
                auc = (sensitivity + specificity) / 2
                
                summary_ws['A1'] = "RESUMEN DE MÉTRICAS - ANÁLISIS DE CÁNCER DE MAMA"
                summary_ws['A1'].font = Font(size=16, bold=True, color="366092")
                summary_ws['A1'].alignment = Alignment(horizontal="center")
                summary_ws.merge_cells('A1:D1')
                
    
    def _excel_to_bytes(self, workbook):
        """Convertir workbook a bytes"""
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
